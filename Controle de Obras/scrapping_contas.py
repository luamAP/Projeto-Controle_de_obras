# %%
from selenium import webdriver

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains

from time import sleep as para
from time import time as contar

from openpyxl import load_workbook, Workbook

# %%
class Navegacao():
    def __init__(self):
        options = webdriver.ChromeOptions()
        if input('Deseja esconder o navegador? y/n  ')=='y': options.add_argument("--headless")
        self.navegador = webdriver.Chrome(options=options)
        self.vars = {'conta_contabil': [],
                    'ignored_exceptions': (NoSuchElementException,StaleElementReferenceException), 
                    'senha': ['Jaspekde27!','Jaspekde27.9', 'Jaspekde279310', 'Luam2722']}

    def esperar_elemento(self, param1, param2, elementos=False, tempo_parar=30):
        """Espera um elemento carregar na página

        Args:
            param1 (): tipo do elemento
            param2 (str): nome do elemento
            tempo0 (int, optional): tempo determinado para executar a espera. Padrão 30 (0,5 min).
        """    
        tempo0 = contar() 
        
        if elementos:
            try:
                elemento =self.navegador.find_elements(param1, param2)
                while ((tempo-tempo0)>=tempo_parar) and (len(elemento)==0):
                    para(1)
                    tempo = contar() 
                else: print(f'Tempo excedido!')

            except:
                elemento = WebDriverWait(self.navegador, tempo_parar,ignored_exceptions=self.vars['ignored_exceptions'])\
                            .until(expected_conditions.presence_of_all_elements_located((param1, param2)))
                tempo = contar()

            finally:
                print(f'>>> Elementos "{param2}" encontrados em {round(tempo-tempo0,2)}s! (len({len(elemento)}))')
                return elemento
        else:
            try:
                elemento =self.navegador.find_elements(param1, param2)
                while ((tempo-tempo0)>=tempo_parar) and (len(elemento)==0):
                    para(1)
                    tempo = contar() 
                else: print(f'Tempo excedido!')
                    
            except:
                elemento = WebDriverWait(self.navegador, tempo_parar, ignored_exceptions=self.vars['ignored_exceptions'])\
                                .until(expected_conditions.presence_of_element_located((param1, param2)))
                tempo = contar()
            
            finally:
                try: print(f'>>> Elemento "{elemento.text}" encontrado em {tempo-tempo0}s!')
                except: print(f'>>> Elemento "{elemento}" encontrado em {tempo-tempo0}s!')
                return elemento

    ### Esperar uma página carregar -> *_esperar_aba(param)_*
    def navegar_aba(self):
        print('Esperando aba/página abrir!')
        paginas = len(self.navegador.window_handles)
        while paginas <= len(self.navegador.window_handles): para(1)
        para(1)
        
        print('Página carregada!')

    def entra_AFIM(self, afim):
        """Abre o AFIM e direciona para a função LISNE

        Args:
            afim (_type_): Ano no qual o FIM irá abrir
        """
        print(f'\n=============== Executando entra_AFIM {afim}! ===============')
        
        if afim in ['2014','2015','2016','2017','2018']:
            #AFIM DE 2018
            if afim=='2018': self.navegador.get('http://afim2.manaus.am.gov.br/AfimPRD2018/')
            #AFIM DE 2017
            elif afim=='2017': self.navegador.get('http://afim2.manaus.am.gov.br/AfimPRD2017/')
            #AFIM DE 2016
            elif afim=='2016': self.navegador.get('http://afim2.manaus.am.gov.br/AfimPRD2016/')
            #AFIM DE 2015
            elif afim=='2015': self.navegador.get('http://afim2.manaus.am.gov.br/AfimPRD2015/')
            #AFIM DE 2014
            elif afim=='2014': self.navegador.get('http://afim2.manaus.am.gov.br/AfimPRD2014/')
        
            self.navegador.find_element(By.NAME, 'user').send_keys('03441042220')
            self.navegador.find_element(By.NAME, 'password').send_keys(self.vars['senha'][3])
            self.navegador.find_element(By.NAME, 'enviar').click()
            para(1)
            self.navegador.switch_to.window(self.navegador.window_handles[-1])
            self.navegador.switch_to.frame(1)

        #AFIM DE 2019
        if afim =='2019':
            self.navegador.get('https://afim1.manaus.am.gov.br/AfimPRD2019/')
            self.navegador.find_element(By.ID, 'user').send_keys('034.410.422-20')
            # Mudar as senhas periodocamente
            self.navegador.find_element(By.ID, 'password').send_keys(self.vars['senha'][3])
            self.navegador.find_element(By.ID, 'password').send_keys(Keys.ENTER)

            para(2)
            for w in self.navegador.window_handles:
                self.navegador.switch_to.window(w)
                if 'mensagem' in self.navegador.current_url: self.navegador.close()
                elif afim in self.navegador.current_url: pass
                else: self.navegador.close()
                
            self.navegador.switch_to.window(self.navegador.window_handles[-1])
            self.navegador.switch_to.frame(1)

        if afim in ['2020', '2021', '2022']:
            #AFIM DE 2022
            if afim=='2022': self.navegador.get('https://afim1.manaus.am.gov.br/AfimPRD2022/')
            #AFIM DE 2021
            elif afim=='2021': self.navegador.get('https://afim1.manaus.am.gov.br/AfimPRD2021/')
            #AFIM DE 2020
            elif afim=='2020': self.navegador.get('https://afim1.manaus.am.gov.br/AfimPRD2020/')
            
            self.navegador.find_element(By.ID, 'user').send_keys('034.410.422-20')
            # Mudar as senhas periodocamente
            self.navegador.find_element(By.ID, 'password').send_keys(self.vars['senha'][3])
            self.navegador.find_element(By.ID, 'password').send_keys(Keys.ENTER)
            self.navegador.switch_to.window(self.navegador.window_handles[-1])
            self.navegador.switch_to.frame(0)

        try: self.navegador.find_element(By.CSS_SELECTOR, "body").click()
        except: self.navegador.esperar_elemento(By.CSS_SELECTOR, "body").click()

        print(f'=============== Finalizando entra_AFIM {afim}! ===============\n')
        ## Mudando de janela,
        para(2)

    def NL_detalhada(self, obj_NL):
        """Pesquisa a Nota de Lançamento desejada para retirar as informações

        Args:
            obj_NL (_type_): Objeto Selenium da Nota de Lançamento

        Returns:
            _type_: Tupla com as informações extraidas
        """
        if not 'NL' in obj_NL.text: 
            print(f'\n=============== Impossível executar! Não é uma NL ({obj_NL.text}) ===============')
            return ('', '', '', '', '', '')
        
        
        print(f'\n=============== Executando a NL ({obj_NL.text}) ===============')
        para(2)
        self.navegador.switch_to.window(self.navegador.window_handles[-1])

        elemento = obj_NL
        ActionChains(self.navegador)\
        .double_click(elemento)\
        .perform()

        para(1)
        self.navegador.switch_to.window(self.navegador.window_handles[-1])
        
        credor_empresa = self.esperar_elemento(By.CSS_SELECTOR, 'tr td input[name="txtCredor"]').get_attribute('value')
        obs = self.esperar_elemento(By.CSS_SELECTOR, 'textarea[name="txtObservacao"]').text
        NE = self.esperar_elemento(By.CSS_SELECTOR, 'input[name="txtNuNe"').get_attribute('value')
        natureza_despesa = self.esperar_elemento(By.CSS_SELECTOR, 'input[name="txtRdNaturezaDespesa"]').get_attribute('value')
        doc = self.esperar_elemento(By.CSS_SELECTOR, 'input[name="txtRdNuDocmento"').get_attribute('value')
        
        self.navegador.close()
        self.navegador.switch_to.window(self.navegador.window_handles[-1])
        
        return (obj_NL.text, credor_empresa, obs, NE, natureza_despesa, doc)

    def RAZAO(self, ano):
        print(f'\n=============== Executando RAZAO {ano}! ===============')
        # Ano anterior a 2020
        if ano >= 2020:
            self.navegador.find_element(By.ID, 'dados').send_keys('RAZAO')
            self.navegador.find_element(By.ID, 'dados').send_keys(Keys.ENTER)
        else:
            self.navegador.find_element(By.CSS_SELECTOR, "body").click()
            self.navegador.find_element(By.XPATH, '/html/body/table/tbody/tr[2]/td/ul/li[7]').click()
            self.navegador.find_element(By.XPATH, '/html/body/table/tbody/tr[2]/td/ul/ul[7]/li[1]').click()
            self.navegador.find_element(By.LINK_TEXT, "RAZAO").click()

        para(1)
        self.navegador.switch_to.window(self.navegador.window_handles[-1])
        self.vars['conta_contabil'] = [cbo.get_attribute('Value') for cbo in self.esperar_elemento(By.CSS_SELECTOR, "select[name='cboConta'] option[value^='1232105']", True, 20)]
        para(1)

    def RAZAO_detalhada(self, ano, conta):          
        print(f'\n=============== Executando RAZAO_detalhada em {ano}! ===============')

        ano = str(ano)
        # Ano anterior a 2020
        if int(ano) >= 2020:
            print(f'Selecionando a Data Inicial: 01/01/{ano}')
            if self.esperar_elemento(By.ID, 'mkdDataInicio').get_attribute('value')!=f'01/01/{ano}':
                self.esperar_elemento(By.ID, 'mkdDataInicio').click()
                self.esperar_elemento(By.CSS_SELECTOR, "li[data-view='month current']").click()
                self.esperar_elemento(By.CSS_SELECTOR, "ul li[data-view='month']", True, 30)[0].click()
            else: print('Data Inicial já selecionada')
                
            print(f'Selecionando a Data Final: 31/12/{ano}')
            if self.esperar_elemento(By.ID, 'mkdDataFim').get_attribute('value')!=f'31/12/{ano}':
                self.esperar_elemento(By.ID, 'mkdDataFim').click()
                self.esperar_elemento(By.CSS_SELECTOR, "div[class*='datepicker-top-right'] li[data-view='month current']").click()
                self.esperar_elemento(By.CSS_SELECTOR, "div[class*='datepicker-top-right'] ul[data-view='months'] li[data-view='month']", True, 30)[-1]
            else: print('Data Final já selecionada')

        # Ano 2020 em diante
        else:
            print(f'Selecionando a Data Inicial: 01/01/{ano}')
            if self.esperar_elemento(By.ID, 'mkdDataInicio').get_attribute('value')!=f'01/01/{ano}':
                self.esperar_elemento(By.ID, 'img0').click()
                while not (ano in self.navegador.find_elements(By.CSS_SELECTOR, "font[size='1']")[0].text): self.navegador.find_element(By.ID, "toLeft").click()
                if self.navegador.find_element(By.CSS_SELECTOR, "#month [selected]").text!='Janeiro': self.esperar_elemento(By.CSS_SELECTOR, "#month").send_keys('Janeiro')
                self.esperar_elemento(By.CSS_SELECTOR, "td[class='calReg']", True, 30)[0].click()
            else: print('Data Inicial já selecionada')

            print(f'Selecionando a Data Final: 31/12/{ano}')
            if self.esperar_elemento(By.ID, 'mkdDataFim').get_attribute('value')!=f'31/12/{ano}':
                self.esperar_elemento(By.ID, 'img1').click()
                while not (ano in self.navegador.find_elements(By.CSS_SELECTOR, "font[size='1']")[0].text): self.navegador.find_element(By.ID, "toLeft").click()
                if self.navegador.find_element(By.CSS_SELECTOR, "#month [selected]").text!='Dezembro': self.esperar_elemento(By.CSS_SELECTOR, "#month").send_keys('Dezembro')
                self.esperar_elemento(By.CSS_SELECTOR, "td[class='calReg']", True, 30)[-1].click()
            else: print('Data Final já selecionada')

        self.navegador.switch_to.window(self.navegador.window_handles[-1])

        self.esperar_elemento( By.CSS_SELECTOR, 'select[name="cboUnidadeGestora"]').send_keys("270101 - Secretaria Municipal de Infraestrutura")
        self.navegador.find_element(By.CSS_SELECTOR, 'select[name="cboUnidadeGestora"]').click()

        self.esperar_elemento(By.NAME, 'cboConta').send_keys(conta)

        self.esperar_elemento(By.ID, 'txtContaCorrente').clear
        self.esperar_elemento(By.ID, 'txtContaCorrente').send_keys('@')

        conta_nominada = self.navegador.find_element(By.CSS_SELECTOR, f"select[name='cboConta'] option[value='{conta}']").text

        para(1)
        try: self.esperar_elemento(By.ID, 'IMGprocurar').click()
        except: self.esperar_elemento(By.ID, 'btnProcurar').click()

        try: 
            pl = load_workbook(f'C:\\Users\\luan.pinto\\Desktop\\Códigos\\Projeto - Análises Simplificadas para o DAF\\Projeto ajustando Excel\\Controle de Obras\\123210500\\123210500-BENS DE USO COMUM DO POVO - {ano}.xlsx')
            print(f'Abrindo o arquivo "C:\\Users\\luan.pinto\\Desktop\\Códigos\\123210500\\123210500-BENS DE USO COMUM DO POVO - {ano}.xlsx"')
        except: 
            pl = Workbook()
            print(f'Criando um arquivo!')

        for i in['/','\\',':']: 
            if i in conta_nominada: conta_nominada = conta_nominada.replace(i,'')

        if conta_nominada in pl: return print(f'Conta já existente na planilha! ({conta_nominada})')

        nova_aba = []
        # Tabela
        for n in self.navegador.find_elements(By.CSS_SELECTOR, '#dbGrid tr'): 
            nova_aba.append(n.text.split('  '))

        nl_aba = [['NOTA DE LANÇAMENTO','CREDOR','OBSERVAÇÃO',
        'NOTA DE EMPENHO','NATUREZA DE DESPESA','N° DO DOCUMENTO']]

        for i in self.navegador.find_elements(By.CSS_SELECTOR, '#dbGrid tr td[onclick="selectCell(this,3)"]'): nl_aba.append(self.NL_detalhada(i))

        # try:
        temp = pl.create_sheet(conta_nominada)
        for indice,linha in enumerate(nova_aba):
            print(f'{indice}: {linha} + {nl_aba[indice-1]}') 
            linha.extend(nl_aba[indice])
            temp.append(linha)
        # except:
        #     return {'nl_aba': nl_aba, 'nova_aba': nova_aba}

        pl.save(f'C:\\Users\\luan.pinto\\Desktop\\Códigos\\Projeto - Análises Simplificadas para o DAF\\Projeto ajustando Excel\\Controle de Obras\\123210500\\123210500-BENS DE USO COMUM DO POVO - {ano}.xlsx')

        print('=============== Finalizando RAZAO! ===============')
        # return f'C:\\Users\\luan.pinto\\Desktop\\Códigos\\Projeto - Análises Simplificadas para o DAF\\Projeto ajustando Excel\\Controle de Obras\\123210500\\123210500-BENS DE USO COMUM DO POVO - {ano}.xlsx'
        

# %%
def linha(wb):
    lista = ['CONTA CONTÁBIL']
    for row in range(wb.max_column):
        lista.append(wb.cell(1, row+1).value)

    return lista

def contas_unidas(ano):
    pl = load_workbook(f'C:\\Users\\luan.pinto\\Desktop\\Códigos\\Projeto - Análises Simplificadas para o DAF\\Projeto ajustando Excel\\Controle de Obras\\123210500\\123210500-BENS DE USO COMUM DO POVO - {ano}.xlsx')
    sheets = pl.sheetnames
    for sheet in sheets:
        print(f'==================== {sheet} ====================')
        aux=[]
        if sheet == 'Sheet': 
            pl[sheet].title = f'123210500-BENS DE USO COMUM DO POVO {ano}'
            aba = pl[f'123210500-BENS DE USO COMUM DO POVO {ano}']
            aba.append(linha(pl[sheets[-1]]))
        else:
            for row in pl[sheet].iter_rows(2,): 
                aux.append(f'{sheet}')
                for i in row: aux.append(i.value)
                aba.append(aux)
                aux.clear()

    pl.save(f'C:\\Users\\luan.pinto\\Desktop\\Códigos\\Projeto - Análises Simplificadas para o DAF\\Projeto ajustando Excel\\Controle de Obras\\123210500\\123210500 - Todas as contas de {ano}.xlsx')

# %%
for ano in range(2014, 2022):
    N = Navegacao()
    N.entra_AFIM(str(ano))
    N.RAZAO(ano)
    for conta in N.vars['conta_contabil']: N.RAZAO_detalhada(ano, conta)
    N.navegador.close()
    contas_unidas(ano)

# %%
# Imprime a tabela

def tabela_print(pl, aba):
    print(aba.max_column, aba.max_row)
    for j in range(1, aba.max_row + 1):
        for i in range(1, aba.max_column + 1):
            print(aba.cell(row = j, column = i).value, end='     ')
        print('') 

    pl.save('tt.xlsx')


